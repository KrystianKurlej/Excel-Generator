import calendar
import openpyxl
import json
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, NamedStyle, Border, Side, PatternFill
from colorama import init, Fore, Back, Style

init()


def is_input_correct(input):
    try:
        int(input)
        return True
    except ValueError:
        return False


def take_input(question):
    while True:
        answer = input(question)
        if is_input_correct(answer):
            return int(answer)
        else:
            print("Podaj poprawną liczbę.")


def get_polish_month_name(month):
    polish_month_names = [
        "Styczeń",
        "Luty",
        "Marzec",
        "Kwiecień",
        "Maj",
        "Czerwiec",
        "Lipiec",
        "Sierpień",
        "Wrzesień",
        "Październik",
        "Listopad",
        "Grudzień",
    ]
    return polish_month_names[month - 1] if 1 <= month <= 12 else ""


def is_weekend(year, month, day):
    return calendar.weekday(year, month, day) in (5, 6)


def create_excel_file(year, month):
    month_name = get_polish_month_name(month)

    # Wczytaj listę osób z pliku JSON
    with open("data/members.json", "r") as file:
        members_list = json.load(file)

    # Utwórz nowy plik Excela i dodaj arkusz
    workbook = Workbook()
    sheet = workbook.active

    # Ustaw nazwę arkusza
    sheet.title = f"{month_name} {year} Obecności"

    # Wprowadź "imię i nazwisko" do komórki A1
    sheet["A1"] = "Imię i nazwisko"

    # Definiuj styl dla komórek
    cell_style = NamedStyle(name="cell_style")
    cell_style.font = Font(bold=True)
    cell_style.border = Border(
        left=Side(border_style="thin"),
        right=Side(border_style="thin"),
        top=Side(border_style="thin"),
        bottom=Side(border_style="thin"),
    )
    cell_style.fill = PatternFill(
        start_color="EFEFEF", end_color="EFEFEF", fill_type="solid"
    )

    # Zastosuj styl do kolumny A
    for cell in sheet["A"]:
        cell.style = cell_style

    # Uzyskaj liczbę dni w danym miesiącu
    num_days_in_month = calendar.monthrange(year, month)[1]

    # Wprowadź dni miesiąca jako liczby od 1 do num_days_in_month w komórkach od B1 w prawo
    column_idx = 2  # Zaczynamy od komórki B1
    for day in range(1, num_days_in_month + 1):
        if not is_weekend(year, month, day):
            sheet.cell(row=1, column=column_idx, value=day).style = cell_style
            column_idx += 1

    # Wprowadź osoby z listy do arkusza, zaczynając od komórki A2
    for row_idx, person in enumerate(members_list, start=2):
        sheet.cell(row=row_idx, column=1, value=person).style = cell_style

    # Dodaj regułę do każdej kolumny, która oblicza sumę komórek z "x"
    for column in sheet.iter_cols(
        min_col=2, max_col=sheet.max_column, min_row=2, max_row=sheet.max_row
    ):
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            cell.value = "x" if cell.value == "x" else None
        last_person_row = len(column) + 1
        sum_formula = (
            f'=COUNTIF({column_letter}2:{column_letter}{last_person_row}, "x")'
        )
        sheet.cell(
            row=last_person_row + 1, column=column[0].column, value=sum_formula
        ).style = cell_style

    # Określ kolumnę, w której powinna znajdować się funkcja sumująca
    sum_column = sheet.max_column + 1

    # Ustawienie szerokości kolumn
    sheet.column_dimensions["A"].width = 20
    for col in range(2, sum_column):
        sheet.column_dimensions[get_column_letter(col)].width = 3
    sheet.column_dimensions[get_column_letter(sum_column)].width = 8

    # Dodaj napis "Suma" w pierwszym wierszu kolumny z sumowaniem
    sheet.cell(row=1, column=sum_column, value="Suma").style = cell_style

    # Dodaj regułę do każdego wiersza, która oblicza sumę komórek z "x" za ostatnim dniem miesiąca
    for row_idx, person in enumerate(members_list, start=2):
        row_formula = f'=COUNTIF(B{row_idx}:{get_column_letter(sheet.max_column - 1)}{row_idx}, "x")'
        sheet.cell(row=row_idx, column=sum_column, value=row_formula).style = cell_style

    # Zapisz arkusz do pliku
    filename = f"Lista obecności - {month_name} {year}.xlsx"
    workbook.save(filename)
    print(
        f"{Fore.GREEN}Utworzono plik: Lista obecności - {month_name} {year}.xlsx{Style.RESET_ALL}"
    )


if __name__ == "__main__":
    year = take_input(Fore.CYAN + "Podaj rok: " + Style.RESET_ALL)
    month = take_input(
        Fore.CYAN + "Podaj miesiąc (w formacie 1-12): " + Style.RESET_ALL
    )

    if month < 1 or month > 12:
        print(Fore.RED + "Podano niepoprawny miesiąc." + Style.RESET_ALL)
    else:
        create_excel_file(year, month)
